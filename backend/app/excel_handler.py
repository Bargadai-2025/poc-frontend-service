import ast
import json
from io import BytesIO
from openpyxl import load_workbook, Workbook
from typing import List, Dict, Any

# Application uses SL (Scoreplex) attribute names for fetching; output Excel uses same names as column headers.
# Order after task_id = exact sequence below (from SL).
DATA_COLUMNS_ORDER = [
    "full_names",           # ✅ NEW
    "full_names_number",
    "email_address_amount",
    "email_addresses",
    # "email_addresses_update",
    "phone_numbers_amount",
    "phone_numbers_list",
    "phone_numbers_list_update",
    "email_valid",
    "email_deliverability",
    "email_disposable",
    "email_first_name",
    "email_phone_numbers",
    "email_generic",
    "email_common",
    "email_spam_trap_score",
    "email_frequent_complainer",
    "email_suspect",
    "email_recent_abuse",
    "email_domain_age",
    "email_domain_velocity",
    "email_domain_trust",
    "email_suggested_domain",
    "email_smtp_score",
    "email_overall_score",
    "email_risky_tld",
    "email_spf_record",
    "email_dmarc_record",
    "email_mx_records",
    "email_mx_record_count",
    "email_format_is_bad",
    "email_has_only_one_digit",
    "email_has_stop_words",
    "email_name_risk_type",
    "email_account_vowels_count",
    "email_account_consonants_count",
    "email_account_length",
    "email_account_digit_count",
    "email_data_leaks_count",
    "email_data_leaks_list",
    "email_data_leaks_records",
    "email_data_leaks_first_seen",
    "email_data_leaks_last_seen",
    "email_string_length",
    "phone_valid",
    "phone_active",
    "phone_associated_emails",
    "phone_name",
    "phone_carrier",
    "phone_line_type",
    "phone_recent_abuse",
    "phone_spammer",
    "phone_voip",
    "phone_prepaid",
    "phone_risky",
    "phone_country",
    "phone_city",
    "phone_region",
    "phone_zip_code",
    "phone_timezone",
    "Phone_No_Frequent_Abuse_History",
    "Phone_No_Reachability",
    "phone_data_leaks_count",
    "phone_data_leaks_list",
    "phone_data_leaks_records",
    "phone_data_leaks_first_seen",
    "phone_data_leaks_last_seen",
]
# Order: email_status, data_leak_status, phone_status, then result_status and result_err
STATUS_COLUMNS_AT_END = ["email_status", "data_leak_status", "phone_status", "result_status", "result_err", "sl_data_phones",       # ✅ NEW
    "sl_data_emails",       # ✅ NEW
    "sl_data_full_names",   # ✅ NEW
    "sl_data_aliases",      # ✅ NEW
    "sl_data_accounts",     # ✅ NEW
    "sl_data_addresses",    # ✅ NEW
    "sl_data_genders",      # ✅ NEW
    "sl_data_birthdays",    # ✅ NEW
]

def _format_list_plain(val: Any) -> str:
    """Format list/JSON array as plain comma-separated (no brackets, no double quotes)."""
    if val is None:
        return ""
    if isinstance(val, list):
        return ", ".join(str(x).strip().strip('"').strip("'") for x in val if x is not None)
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("["):
            try:
                arr = json.loads(s)
                return ", ".join(str(x).strip().strip('"').strip("'") for x in arr if x is not None)
            except Exception:
                return s.strip("[]").replace('"', "").replace("'", "")
        return s.replace('"', "").replace("'", "")
    return str(val)


def _format_data_leaks_records(val: Any) -> str:
    """Format data leaks list of dicts as '(Title: X, Breach Date: Y); ' — each item in parentheses."""
    if val is None:
        return ""
    items = []
    if isinstance(val, str):
        try:
            val = json.loads(val)
        except Exception:
            return val
    if not isinstance(val, list):
        return str(val)
    for item in val:
        if isinstance(item, dict):
            title = item.get("title") or item.get("Title") or item.get("name") or ""
            breach = item.get("breach_date") or item.get("Breach Date") or item.get("breach_date_date") or ""
            items.append(f"(Title: {title}, Breach Date: {breach});")
        else:
            items.append(f"({str(item)});")
    return " ".join(items)


def _normalize_key(k: str) -> str:
    return (k or "").strip().replace("  ", " ")


def _is_empty(val: Any) -> bool:
    """Treat None and blank string as empty; 0 and False are valid (non-empty)."""
    if val is None:
        return True
    if isinstance(val, str) and str(val).strip() == "":
        return True
    return False


def _camel_to_snake(name: str) -> str:
    """Convert camelCase or mixed to snake_case for matching Scoreplex keys to SL columns."""
    if not name:
        return ""
    s = str(name).strip()
    out = []
    for i, c in enumerate(s):
        if c.isupper() and i > 0 and (i + 1 >= len(s) or not s[i + 1].isupper()):
            out.append("_")
        out.append(c.lower())
    return "".join(out).replace(" ", "_").replace("-", "_").replace("__", "_").strip("_")


def _is_valid_phone_for_output(phone: Any) -> bool:
    """True if string looks like a 10-digit number or 12-digit with 91 (India); exclude other alternates."""
    if phone is None or (isinstance(phone, str) and not phone.strip()):
        return False
    s = str(phone).strip()
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 10:
        return True
    if len(digits) == 12 and digits.startswith("91"):
        return True
    if len(digits) == 11 and digits.startswith("91"):
        return True
    return False


def _normalize_phone_91(phone: Any) -> str:
    """Ensure phone in output has 91 prefix then 10 digits (91 + 10 = 12 digits)."""
    if phone is None:
        return ""
    s = str(phone).strip()
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return s
    if digits.startswith("91") and len(digits) >= 12:
        return digits[:12]
    if digits.startswith("91") and len(digits) == 11:
        return "91" + digits[2:].zfill(10) if len(digits[2:]) <= 10 else digits[:12]
    if len(digits) == 10:
        return "91" + digits
    if len(digits) > 10 and not digits.startswith("91"):
        return "91" + digits[-10:]
    return "91" + digits[-10:] if len(digits) >= 10 else "91" + digits.zfill(10)


def _flask_email_mx_record_count(mx_records: Any) -> Any:
    """Return total count of items in email_mx_records (list, dict, or JSON string)."""
    if mx_records is None:
        return 0
    if isinstance(mx_records, list):
        return len(mx_records)
    if isinstance(mx_records, dict):
        return len(mx_records)
    if isinstance(mx_records, str):
        s = mx_records.strip()
        if not s:
            return 0
        try:
            parsed = json.loads(s)
            return len(parsed) if isinstance(parsed, (list, dict)) else (1 if parsed else 0)
        except (json.JSONDecodeError, TypeError):
            pass
        try:
            parsed = ast.literal_eval(s)
            return len(parsed) if isinstance(parsed, (list, dict)) else (1 if parsed else 0)
        except (ValueError, SyntaxError, TypeError):
            pass
        if s.startswith("[") and "]" in s:
            return max(0, s.count(",") + 1)
            return len([x for x in s.split(",") if x.strip()])
    return 0


def _flask_email_string_length(account_length: Any, digit_count: Any) -> Any:
    """Flask mapping: email_string_length = email_account_length - email_account_digit_count (max 0)."""
    try:
        al = None if account_length is None or account_length == "" else float(account_length)
        dc = None if digit_count is None or digit_count == "" else float(digit_count)
    except (ValueError, TypeError):
        return ""
    if al is not None and dc is not None:
        return max(0, int(al - dc))
    return ""


def _flask_phone_reachability(phone_valid: Any, phone_active: Any) -> bool:
    """Flask mapping: Phone_No_Reachability = phone_valid AND phone_active (boolean)."""
    def to_bool(val: Any) -> bool:
        if val is None or val == "":
            return False
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            return bool(val)
        if isinstance(val, str):
            v = val.lower().strip()
            if v in ("true", "1", "yes", "y"):
                return True
            if v in ("false", "0", "no", "n", ""):
                return False
            return bool(val)
        return bool(val)
    return to_bool(phone_valid) and to_bool(phone_active)


def _normalize_flat_key(k: str) -> str:
    """Strip report_ prefix and lowercase for flexible lookup."""
    s = (k or "").strip().lower()
    for prefix in ("report_", "report-"):
        if s.startswith(prefix):
            s = s[len(prefix) :].lstrip("_-")
            break
    return s.replace("  ", " ")


def _sl_attr_domain(sl_attr: str) -> str:
    """Return 'email', 'phone', or 'ip' based on sl_attr prefix."""
    a = sl_attr.lower()
    if a.startswith("email_"):
        return "email"
    if a.startswith("phone_"):
        return "phone"
    if a.startswith("ip_"):
        return "ip"
    return ""


# Flask-style keyword fallback: SL column -> keywords that may appear in API key (with domain check)
_KEYWORD_MAP = {
    "email_address_amount": ["address_amount", "addressamount", "address", "amount"],
    "email_disposable": ["disposable"],
    "email_generic": ["generic"],
    "email_frequent_complainer": ["frequent_complainer", "frequentcomplainer", "complainer", "frequent"],
    "email_suspect": ["suspect"],
    "email_recent_abuse": ["recent_abuse", "recentabuse", "recent", "abuse"],
    "email_risky_tld": ["risky_tld", "riskytld", "risky", "tld"],
    "email_format_is_bad": ["format_is_bad", "formatisbad", "format", "bad"],
    "email_has_only_one_digit": ["has_only_one_digit", "hasonlyonedigit", "one_digit", "digit"],
    "email_has_stop_words": ["has_stop_words", "hasstopwords", "stop_words", "stopwords"],
    "email_data_leaks_list": ["data_leaks_list", "dataleakslist", "data_leaks", "leaks", "list"],
    "email_data_leaks_records": ["data_leaks_records", "dataleaksrecords", "data_leaks", "records"],
    "email_string_length": ["string_length", "stringlength", "string", "length"],
    "phone_associated_emails": ["associated_emails", "associatedemails", "associated", "emails"],
    "phone_spammer": ["spammer"],
    "phone_recent_abuse": ["recent_abuse", "recentabuse", "recent", "abuse"],
    "phone_voip": ["voip"],
    "phone_prepaid": ["prepaid"],
    "phone_risky": ["risky"],
    "phone_data_leaks_count": ["data_leaks_count", "dataleakscount", "data_leaks", "leaks", "count"],
    "phone_data_leaks_list": ["data_leaks_list", "dataleakslist", "data_leaks", "leaks", "list"],
    "phone_data_leaks_records": ["data_leaks_records", "dataleaksrecords", "data_leaks", "records"],
    "Phone_No_Frequent_Abuse_History": ["no_frequent_abuse_history", "nofrequentabusehistory", "frequent_abuse", "abuse_history", "frequent", "abuse"],
    "Phone_No_Reachability": ["no_reachability", "noreachability", "reachability", "reachable"],
}

# Alternate status key names (match Scoreplex / Flask - report=false may use camelCase)
_STATUS_KEYS = {
    "email_status": ("email_status", "emailStatus", "status_email", "email"),
    "phone_status": ("phone_status", "phoneStatus", "status_phone", "phone"),
    "data_leak_status": ("data_leak_status", "dataLeakStatus", "data_leaks_status", "leaks_status"),
}


def _get_status_from_data(data: Dict[str, Any], key_list: tuple) -> Any:
    """Return first value found for any of the keys (case-insensitive)."""
    if not data:
        return None
    data_lower = {str(k).lower(): v for k, v in data.items()}
    for key in key_list:
        if key.lower() in data_lower:
            return data_lower[key.lower()]
    return None


def _key_domain(key: str) -> str:
    """Infer domain from a flat key (e.g. report_email_valid -> email)."""
    k = key.lower().replace(".", "_").replace("-", "_")
    if "email_" in k or k.startswith("email_"):
        return "email"
    if "phone_" in k or k.startswith("phone_"):
        return "phone"
    if "ip_" in k or k.startswith("ip_"):
        return "ip"
    return ""


def _extract_response_data(response: Any) -> Any:
    """Unwrap nested API response like Flask: try data, result, response, body, content, payload, report; recurse until no further nesting."""
    if not isinstance(response, dict):
        return response
    nested_keys = ["data", "result", "response", "body", "content", "payload", "report"]
    for key in nested_keys:
        if key not in response:
            continue
        val = response[key]
        if isinstance(val, dict):
            if any(nk in val and isinstance(val.get(nk), dict) for nk in nested_keys):
                return _extract_response_data(val)
            return val
        if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
            return val[0]
    return response


def _find_value_in_flat(flat: Dict[str, Any], sl_attr: str) -> Any:
    """Get value from flattened report using multi-strategy matching; prefer non-empty for consistent results.
    Matches Scoreplex camelCase keys (e.g. emailValid) to SL column names (e.g. email_valid) like Flask."""
    sl_lower = sl_attr.lower()
    if sl_attr in flat and not _is_empty(flat[sl_attr]):
        return flat[sl_attr]
    # Match by camelCase -> snake_case (Scoreplex returns camelCase; SL columns are snake_case - same as Flask)
    for k, v in flat.items():
        if _camel_to_snake(k).lower() == sl_lower and not _is_empty(v):
            return v
    for k, v in flat.items():
        if _camel_to_snake(k).lower() == sl_lower:
            return v
    n = _normalize_key(sl_attr).lower()
    domain = _sl_attr_domain(sl_attr)
    # 1) Normalized key match (prefer non-empty)
    for k, v in flat.items():
        if _normalize_flat_key(k) == n and not _is_empty(v):
            return v
    for k, v in flat.items():
        if _normalize_flat_key(k) == n:
            return v
    # 2) Available key ends with sl_attr (e.g. report_email_valid -> email_valid); prefer non-empty
    ends_match = []
    for k, v in flat.items():
        k_lower = k.lower()
        if k_lower.endswith("_" + sl_lower) or k_lower == sl_lower:
            k_domain = _key_domain(k)
            ends_match.append((k_domain == domain, _is_empty(v), k, v))
    if ends_match:
        ends_match.sort(key=lambda x: (not x[0], x[1], x[2]))
        return ends_match[0][3]
    # 3) Available key contains sl_attr as segment (word boundary); prefer non-empty
    segment_match = []
    for k, v in flat.items():
        k_lower = k.lower()
        if ("_" + sl_lower + "_" in k_lower or
                k_lower.startswith(sl_lower + "_") or
                k_lower.endswith("_" + sl_lower)):
            k_domain = _key_domain(k)
            segment_match.append((k_domain == domain, _is_empty(v), k, v))
    if segment_match:
        segment_match.sort(key=lambda x: (not x[0], x[1], x[2]))
        return segment_match[0][3]
    # 4) Flask-style keyword/contains fallback with domain validation
    sl_attr_lower = sl_attr.lower()
    keyword_list = _KEYWORD_MAP.get(sl_attr) or _KEYWORD_MAP.get(sl_attr_lower)
    if not keyword_list:
        for mk, mv in _KEYWORD_MAP.items():
            if mk.lower() == sl_attr_lower:
                keyword_list = mv
                break
    if keyword_list and domain:
        for k, v in flat.items():
            k_lower = k.lower()
            k_dom = _key_domain(k)
            if k_dom != domain:
                continue
            if sl_attr_lower == "email_data_leaks_records" and "mx_records" in k_lower:
                continue
            if sl_attr_lower == "phone_data_leaks_records" and ("mx_records" in k_lower or "email" in k_lower):
                continue
            if sl_attr_lower == "email_string_length" and "string" not in k_lower and "account_length" in k_lower:
                continue
            if sl_attr_lower == "phone_no_frequent_abuse_history" and "employment" in k_lower:
                continue
            for kw in keyword_list:
                if kw.lower() in k_lower:
                    if not _is_empty(v):
                        return v
        for k, v in flat.items():
            k_lower = k.lower()
            k_dom = _key_domain(k)
            if k_dom != domain:
                continue
            if sl_attr_lower == "email_data_leaks_records" and "mx_records" in k_lower:
                continue
            if sl_attr_lower == "phone_data_leaks_records" and ("mx_records" in k_lower or "email" in k_lower):
                continue
            if sl_attr_lower == "email_string_length" and "string" not in k_lower and "account_length" in k_lower:
                continue
            if sl_attr_lower == "phone_no_frequent_abuse_history" and "employment" in k_lower:
                continue
            for kw in keyword_list:
                if kw.lower() in k_lower:
                    return v
    if keyword_list and not domain:
        for k, v in flat.items():
            k_lower = k.lower()
            for kw in keyword_list:
                if kw.lower() in k_lower and not _is_empty(v):
                    return v
        for k, v in flat.items():
            k_lower = k.lower()
            for kw in keyword_list:
                if kw.lower() in k_lower:
                    return v
    return None


class ExcelHandler:
    @staticmethod
    def read_input_excel_from_bytes(content: bytes) -> List[Dict[str, str]]:
        """
        Read input Excel from bytes (in-memory). Email and phone columns required.
        """
        try:
            wb = load_workbook(BytesIO(content), read_only=False)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value.strip().lower() if cell.value else "" 
                      for cell in ws[1]]
            
            # Find email and phone column indices
            try:
                email_idx = headers.index('email')
                phone_idx = headers.index('phone')
            except ValueError:
                raise ValueError("Excel must have 'email' and 'phone' columns")
            # Optional IP column
            ip_idx = None
            if 'ip' in headers:
                ip_idx = headers.index('ip')
            
            # Read data rows (skip header row)
            rows = []
            _HEADER_EMAIL = frozenset(('email', 'e-mail', 'email address', 'email id'))
            _HEADER_PHONE = frozenset(('phone', 'phone number', 'mobile', 'contact', 'contact number', 'phone no', 'phone no.'))
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not row or not any(row):
                    continue
                
                email = str(row[email_idx] or "").strip()
                phone = str(row[phone_idx] or "").strip()
                ip_val = str(row[ip_idx] or "").strip() if ip_idx is not None else ""
                
                # Skip if both email and phone are empty
                if not email and not phone:
                    continue
                
                # Never treat title/header row as data (email and phone column titles)
                email_lower = email.lower()
                phone_lower = phone.lower()
                if email_lower in _HEADER_EMAIL or phone_lower in _HEADER_PHONE:
                    continue
                if (email_lower, phone_lower) == ('email', 'phone'):
                    continue
                
                rows.append({
                    "email": email,
                    "phone": phone,
                    "ip": ip_val
                })
            
            print(f"✅ Read {len(rows)} rows from Excel (excluding header)", flush=True)
            return rows
            
        except Exception as e:
            print(f"❌ Error reading Excel: {str(e)}", flush=True)
            raise
    
    @staticmethod
    def flatten_response(raw_response: Dict[Any, Any]) -> Dict[str, Any]:
        """Extract company-mapped attributes; supports report=true (nested report) or report=false (flat).
        Status columns use alternate keys (dataLeakStatus etc.) so data_leak_status always shows."""
        out = {}
        if not isinstance(raw_response, dict):
            return out
        # Unwrap like Flask (data, result, response, body, content, payload, report) so we flatten the same payload
        report = _extract_response_data(raw_response)
        if not isinstance(report, dict):
            report = raw_response.get("report", raw_response) if isinstance(raw_response, dict) else raw_response
        if not isinstance(report, dict):
            return out

        _PRIORITY_LIST_COLUMNS = ("email_data_leaks_list", "phone_data_leaks_list", "phone_associated_emails")

        def flatten_dict(d: Dict, parent_key: str = "") -> Dict[str, Any]:
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}_{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key).items())
                elif isinstance(v, list):
                    # Preserve list-of-dicts as JSON for priority columns (Flask-style) so _format_data_leaks_records can parse
                    new_key_snake = _camel_to_snake(new_key).lower()
                    new_key_norm = _normalize_flat_key(new_key)
                    is_priority = any(
                        new_key_snake == p or new_key_norm == p.replace("_", "")
                        or p in new_key_snake or p in new_key_norm
                        for p in _PRIORITY_LIST_COLUMNS
                    )
                    if is_priority and v:
                        items.append((new_key, json.dumps(v) if isinstance(v, list) else v))
                    else:
                        items.append((new_key, ", ".join(map(str, v)) if v else ""))
                else:
                    items.append((new_key, v))
            return dict(items)

        flat = flatten_dict(report)
        # Merge nested email/phone/ip into flat (Flask-style)
        for nest_key in ("email", "phone", "ip"):
            if nest_key in report and isinstance(report[nest_key], dict):
                for nk, nv in flatten_dict(report[nest_key], nest_key).items():
                    if nk not in flat:
                        flat[nk] = nv
        # Normalized keys for lookup (raw keys from API + normalized for matching; no snake_case alias to avoid order-dependent fluctuation)
        for k, v in list(flat.items()):
            norm = _normalize_flat_key(k)
            if norm and norm not in flat:
                flat[norm] = v
        # Aliases: API may return slightly different names; include phone columns that sometimes miss (row variance)
        _SL_ALIASES = {
            "email_address_amount": ("email_address_amount", "email_addresses_amount"),
            "Phone_No_Frequent_Abuse_History": (
                "Phone_No_Frequent_Abuse_History",
                "phone_no_frequent_abuse_history",
                "phoneNoFrequentAbuseHistory",
                "no_frequent_abuse_history",
                "noFrequentAbuseHistory",
            ),
            "Phone_No_Reachability": (
                "Phone_No_Reachability",
                "phone_no_reachability",
                "phoneNoReachability",
                "phone_noReachability",
                "no_reachability",
                "noReachability",
                "reachability",
                "noreachability",
            ),
            "phone_prepaid": ("phone_prepaid", "phonePrepaid"),
            "phone_risky": ("phone_risky", "phoneRisky"),
            "phone_recent_abuse": (
                "phone_recent_abuse",
                "phoneRecentAbuse",
                "recent_abuse",
                "recentAbuse",
            ),
        }
        for col in DATA_COLUMNS_ORDER:
            val = None if _is_empty(flat.get(col)) else flat.get(col)
            if _is_empty(val):
                val = _find_value_in_flat(flat, col)
            if _is_empty(val) and col in _SL_ALIASES:
                for alt in _SL_ALIASES[col]:
                    v = flat.get(alt) or _find_value_in_flat(flat, alt)
                    if not _is_empty(v):
                        val = v
                        break
            if not _is_empty(val):
                out[col] = val
        # Reverse pass: map raw Scoreplex keys to SL column when column still empty; prefer non-empty (do not fill with empty)
        _REACHABILITY_KEYS = (
            "no_reachability", "noreachability", "phone_reachability", "phone_no_reachability",
            "phononoreachability", "phone_noreachability", "reachability",
        )
        for flat_key, flat_val in flat.items():
            # Include boolean False for reachability (API returns true/false)
            if _is_empty(flat_val) and flat_val is not False and flat_val is not True:
                continue
            snake = _camel_to_snake(flat_key)
            snake_alt = _normalize_flat_key(flat_key)
            # Explicit: any reachability-like key -> Phone_No_Reachability (Flask shows true/false)
            if out.get("Phone_No_Reachability") in (None, ""):
                if snake.lower() in _REACHABILITY_KEYS or snake_alt in _REACHABILITY_KEYS:
                    out["Phone_No_Reachability"] = flat_val
            for col in DATA_COLUMNS_ORDER:
                if _is_empty(out.get(col)):
                    col_norm = _normalize_flat_key(col)
                    if snake == col or snake_alt == col or snake_alt == col_norm or _normalize_flat_key(flat_key) == col_norm:
                        out[col] = flat_val
                        break
                    if col.lower() == snake or col.lower() == snake_alt:
                        out[col] = flat_val
                        break
        # Final fallback: Phone_No_Reachability from any flat key containing "reachability" (Scoreplex returns true/false)
        if out.get("Phone_No_Reachability") in (None, ""):
            for k, v in flat.items():
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                key_lower = (k or "").lower().replace("-", "_")
                key_snake = _camel_to_snake(k).lower()
                if "reachability" in key_lower or "reachability" in key_snake or "noreachability" in key_lower or "no_reachability" in key_snake:
                    out["Phone_No_Reachability"] = v
                    break
        # email_mx_record_count = count of MX records; ensure email_mx_records is in out so per-row count works
        mx_records = out.get("email_mx_records") or flat.get("email_mx_records") or _find_value_in_flat(flat, "email_mx_records")
        if mx_records is not None:
            if out.get("email_mx_records") is None:
                out["email_mx_records"] = mx_records
            if out.get("email_mx_record_count") in (None, ""):
                if isinstance(mx_records, list):
                    out["email_mx_record_count"] = len(mx_records)
                elif isinstance(mx_records, str):
                    s = mx_records.strip()
                    if s.startswith("["):
                        try:
                            arr = json.loads(s)
                            out["email_mx_record_count"] = len(arr) if isinstance(arr, list) else 1
                        except Exception:
                            out["email_mx_record_count"] = max(0, s.count(",") + 1)
                    else:
                        out["email_mx_record_count"] = len([x for x in s.split(",") if x.strip()])
                else:
                    out["email_mx_record_count"] = mx_records

                            # ✅ Fix email_address_amount = count of email_addresses
        email_addrs = out.get("email_addresses") or flat.get("email_addresses") or _find_value_in_flat(flat, "email_addresses")
        if email_addrs is not None:
            if isinstance(email_addrs, list):
                out["email_address_amount"] = len(email_addrs)
            elif isinstance(email_addrs, str):
                s = email_addrs.strip()
                if not s:
                    out["email_address_amount"] = 0
                elif s.startswith("["):
                    try:
                        out["email_address_amount"] = len(json.loads(s))
                    except Exception:
                        out["email_address_amount"] = len([x for x in s.split(",") if x.strip()])
                else:
                    out["email_address_amount"] = len([x for x in s.split(",") if x.strip()])

        # Special formatting: *_update = list without brackets/quotes
        email_addrs = out.get("email_addresses") or flat.get("email_addresses") or _find_value_in_flat(flat, "email_addresses")
        out["email_addresses_update"] = _format_list_plain(email_addrs) if email_addrs is not None else out.get("email_addresses_update", "")
        phone_list = out.get("phone_numbers_list") or flat.get("phone_numbers_list") or _find_value_in_flat(flat, "phone_numbers_list")
        if phone_list is not None:
            plain = _format_list_plain(phone_list)
            if plain:
                parts = [p.strip() for p in plain.split(",") if p and p.strip()]
                valid_parts = [p for p in parts if _is_valid_phone_for_output(p)]
                out["phone_numbers_amount"] = len(valid_parts)
                out["phone_numbers_list"] = ", ".join(valid_parts) if valid_parts else ""
                out["phone_numbers_list_update"] = ", ".join(_normalize_phone_91(p) for p in valid_parts)
            else:
                out["phone_numbers_amount"] = 0
                out["phone_numbers_list"] = ""
                out["phone_numbers_list_update"] = ""
        else:
            out["phone_numbers_list_update"] = out.get("phone_numbers_list_update", "")
        # email_data_leaks_records = Title and Breach Date only
        dl_val = out.get("email_data_leaks_list") or out.get("email_data_leaks") or flat.get("email_data_leaks_list") or _find_value_in_flat(flat, "email_data_leaks_list")
        if dl_val is not None:
            out["email_data_leaks_records"] = _format_data_leaks_records(dl_val)
        # phone_data_leaks_records = same format
        pdl_val = out.get("phone_data_leaks_list") or flat.get("phone_data_leaks_list") or _find_value_in_flat(flat, "phone_data_leaks_list")
        if pdl_val is not None:
            out["phone_data_leaks_records"] = _format_data_leaks_records(pdl_val)
        # Status columns: try alternate keys (dataLeakStatus etc.)
        for status_key, key_list in _STATUS_KEYS.items():
            val = report.get(status_key)
            if val is None:
                val = _get_status_from_data(report, key_list)
            if val is None:
                val = _get_status_from_data(flat, key_list)
            if val is not None:
                out[status_key] = val

        _SL_DATA_COLS = [
            "sl_data_phones", "sl_data_emails", "sl_data_full_names",
            "sl_data_aliases", "sl_data_accounts", "sl_data_addresses",
            "sl_data_genders", "sl_data_birthdays",
        ]
        for sl_col in _SL_DATA_COLS:
            val = flat.get(sl_col) or report.get(sl_col)
            if val is not None:
                out[sl_col] = val    
        return out
    
    @staticmethod
    def write_output_excel(results: List[Dict[Any, Any]], output_path: str):
        """Write results to Excel: company attributes (snake_case) then status columns at the end."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            # Order: input cols, task_id, then SL columns in sequence, then status at end
            columns = (
                ["input_email", "input_phone", "task_id"]
                + list(DATA_COLUMNS_ORDER)
                + list(STATUS_COLUMNS_AT_END)
            )
            ws.append(columns)
            _NO_DASH_COLS = {"input_email", "input_phone", "task_id", "result_status", "result_err"}

            for result in results:
                # Use only the email/phone we submitted for this row (never from API response)
                submitted_email = result.get("email", "")
                submitted_phone = _normalize_phone_91(result.get("phone", ""))
                row_data = {
                    "input_email": submitted_email,
                    "input_phone": submitted_phone,
                    "task_id": result.get("task_id", ""),
                }
                if result.get("status") in ("SUCCESS", "INCOMPLETE") and "raw_response" in result:
                    row_data.update(ExcelHandler.flatten_response(result["raw_response"]))
                # Flask mapping: overwrite these three with exact Flask calculations
                mx_val = row_data.get("email_mx_records") or row_data.get("emailMxRecords") or row_data.get("mx_records")
                mx_count = _flask_email_mx_record_count(mx_val)
                if mx_count == 0 and row_data.get("email_mx_record_count") not in (None, ""):
                    try:
                        existing = row_data["email_mx_record_count"]
                        if isinstance(existing, (int, float)) and int(existing) >= 0:
                            mx_count = int(existing)
                    except (TypeError, ValueError):
                        pass
                row_data["email_mx_record_count"] = mx_count
                # ✅ Fix email_address_amount
                email_addr_val = row_data.get("email_addresses") or ""
                if email_addr_val:
                    if isinstance(email_addr_val, list):
                        row_data["email_address_amount"] = len(email_addr_val)
                    else:
                        s = str(email_addr_val).strip()
                        if s.startswith("["):
                            try:
                                row_data["email_address_amount"] = len(json.loads(s))
                            except Exception:
                                row_data["email_address_amount"] = len([x for x in s.split(",") if x.strip()])
                        else:
                            row_data["email_address_amount"] = len([x for x in s.split(",") if x.strip()])

                _sl_calc = _flask_email_string_length(row_data.get("email_account_length"), row_data.get("email_account_digit_count"))
                row_data["email_string_length"] = _sl_calc if _sl_calc != "" else row_data.get("email_string_length", "")
                row_data["Phone_No_Reachability"] = _flask_phone_reachability(row_data.get("phone_valid"), row_data.get("phone_active"))
                # Keep input columns exactly as submitted (API response must not overwrite)
                row_data["input_email"] = submitted_email
                row_data["input_phone"] = _normalize_phone_91(result.get("phone", ""))
                for c in DATA_COLUMNS_ORDER:
                    if c not in row_data:
                        row_data[c] = ""
                row_data["email_status"] = row_data.get("email_status", "")
                row_data["phone_status"] = row_data.get("phone_status", "")
                row_data["data_leak_status"] = row_data.get("data_leak_status", "")
                row_data["result_status"] = result.get("status", "")
                row_data["result_err"] = result.get("error", "")
                row = [
                    row_data.get(col, "") if col in _NO_DASH_COLS
                    else ("-" if row_data.get(col, "") in (None, "", [], "None") else row_data.get(col, ""))
                    for col in columns
                ]
                ws.append(row)

            wb.save(output_path)
            print(f"✅ Wrote {len(results)} rows to {output_path}", flush=True)
        except Exception as e:
            print(f"❌ Error writing Excel: {str(e)}", flush=True)
            raise

    @staticmethod
    def write_output_excel_to_bytes(results: List[Dict[Any, Any]]) -> bytes:
        """Build result Excel in memory and return as bytes (no file stored)."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            columns = (
                ["input_email", "input_phone", "task_id"]
                + list(DATA_COLUMNS_ORDER)
                + list(STATUS_COLUMNS_AT_END)
            )
            ws.append(columns)
            _NO_DASH_COLS = {"input_email", "input_phone", "task_id", "result_status", "result_err"}
            for result in results:
                submitted_email = result.get("email", "")
                submitted_phone = _normalize_phone_91(result.get("phone", ""))
                row_data = {
                    "input_email": submitted_email,
                    "input_phone": submitted_phone,
                    "task_id": result.get("task_id", ""),
                }
                if result.get("status") in ("SUCCESS", "INCOMPLETE") and "raw_response" in result:
                    row_data.update(ExcelHandler.flatten_response(result["raw_response"]))
                # Flask mapping: overwrite these three with exact Flask calculations
                mx_val = row_data.get("email_mx_records") or row_data.get("emailMxRecords") or row_data.get("mx_records")
                mx_count = _flask_email_mx_record_count(mx_val)
                if mx_count == 0 and row_data.get("email_mx_record_count") not in (None, ""):
                    try:
                        existing = row_data["email_mx_record_count"]
                        if isinstance(existing, (int, float)) and int(existing) >= 0:
                            mx_count = int(existing)
                    except (TypeError, ValueError):
                        pass
                row_data["email_mx_record_count"] = mx_count
                # ✅ Fix email_address_amount
                email_addr_val = row_data.get("email_addresses") or ""
                if email_addr_val:
                    if isinstance(email_addr_val, list):
                        row_data["email_address_amount"] = len(email_addr_val)
                    else:
                        s = str(email_addr_val).strip()
                        if s.startswith("["):
                            try:
                                row_data["email_address_amount"] = len(json.loads(s))
                            except Exception:
                                row_data["email_address_amount"] = len([x for x in s.split(",") if x.strip()])
                        else:
                            row_data["email_address_amount"] = len([x for x in s.split(",") if x.strip()])

                _sl_calc = _flask_email_string_length(row_data.get("email_account_length"), row_data.get("email_account_digit_count"))
                row_data["email_string_length"] = _sl_calc if _sl_calc != "" else row_data.get("email_string_length", "")
                row_data["Phone_No_Reachability"] = _flask_phone_reachability(row_data.get("phone_valid"), row_data.get("phone_active"))
                row_data["input_email"] = submitted_email
                row_data["input_phone"] = _normalize_phone_91(result.get("phone", ""))
                for c in DATA_COLUMNS_ORDER:
                    if c not in row_data:
                        row_data[c] = ""
                row_data["email_status"] = row_data.get("email_status", "")
                row_data["phone_status"] = row_data.get("phone_status", "")
                row_data["data_leak_status"] = row_data.get("data_leak_status", "")
                row_data["result_status"] = result.get("status", "")
                row_data["result_err"] = result.get("error", "")
                row = [
                    row_data.get(col, "") if col in _NO_DASH_COLS
                    else ("-" if row_data.get(col, "") in (None, "", [], "None") else row_data.get(col, ""))
                    for col in columns
                ]
                ws.append(row)
            buffer = BytesIO()
            wb.save(buffer)
            print(f"✅ Built result Excel in memory ({len(results)} rows)", flush=True)
            return buffer.getvalue()
        except Exception as e:
            print(f"❌ Error building Excel: {str(e)}", flush=True)
            raise
